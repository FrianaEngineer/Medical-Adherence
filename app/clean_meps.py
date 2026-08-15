"""Build and export MEPS adherence data for any year (2020-2023).

Single public entry points:

    from clean_meps import build, run_exports, write_log

    patient_drug, bridge, log = build(2023)   # or 2020, 2021, 2022
    run_exports(2023)             # full notebook-style tables + graphs -> output/2023/
    write_log(log)                # appends a run block to decisions_log.md

``patient_drug`` matches the frame Friana's ``YYYY_clean.ipynb`` produces at
the end of cell 56: one row per (DUPERSID, DRUGIDX) after the CLNK join, the
is_chronic ICD filter, and the chronic-drug filter — with the PSTATS-based
reference-days denominator and PDC-style ``meps_adherence_ratio`` already
computed. RXDAYSUP is summed on deduplicated fills so a single fill linked to
multiple chronic conditions is not double-counted. Multi-condition context
is preserved via ``n_chronic_conditions`` + ``chronic_conditions`` (comma-
joined ICD list) columns and via the ``bridge`` frame.

``bridge`` is one row per (DUPERSID, DRUGIDX, ICD10CDX, CONDIDX) — the
patient-drug × chronic-condition mapping. Use it for condition-level rollups
(never sum days-supply through the patient_drug ICD alias — that column is
only the *primary* ICD encountered). Carries no days-supply so it cannot
inflate totals if summed by accident.

``log`` is a plain dict recording every stage's row/patient counts,
columns dropped, and decisions taken, ready to render in a Streamlit sidebar
or persist to ``decisions_log.md``.

``export_merged_all_years()`` stacks each year's chronic-drug export into
``output/all_years/tables/`` (also refreshed at the end of ``run_exports``).

Column set is the lean notebook export plus paper enrichments used by the
app: ``MARRYXX`` / ``REGIONXX`` (newest→oldest round backfill), ``EDUCYR``,
``RACETHX``, and medication_freq / dose / units / freq_bin from Rx fills.
Sidebar filters use the demography fields; medication_* feed the models tab.
"""

from __future__ import annotations

import calendar
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# Config: what file is what, per year
# ---------------------------------------------------------------------------

# AHRQ ships a fresh set of file numbers per year. Only h248a → h220a/h229a/
# h239a substitution; conditions and person-year files also renumber. Person
# file for 2020 is capitalized as ``H224.xlsx`` on AHRQ's release — matters
# on Linux (case-sensitive), harmless on macOS.
YEAR_FILES = {
    2020: {"rx": "h220a.xlsx", "clnk": "h220if1.xlsx",
           "cond": "h222.xlsx", "person": "H224.xlsx"},
    2021: {"rx": "h229a.xlsx", "clnk": "h229if1.xlsx",
           "cond": "h231.xlsx", "person": "h233.xlsx"},
    2022: {"rx": "h239a.xlsx", "clnk": "h239if1.xlsx",
           "cond": "h241.xlsx", "person": "h243.xlsx"},
    2023: {"rx": "h248a.xlsx", "clnk": "h248if1.xlsx",
           "cond": "h249.xlsx", "person": "h251.xlsx"},
}

# Cross-year lookup tables (same file used for all years).
CHRONIC_ICD_FILE = "is_chronic.xlsx"                          # Friana's DRAFT ICD list
CHRONIC_DRUG_FILE = "unique_rxname_chronic_labeled_revised.csv"  # Friana's labeled RXNAMEs

# PSTATS taxonomy per h251doc.pdf Tables 7-8. Stable across 2020-2023 (the
# 2020 file just has more -1 nonresponse because of COVID).
#
# Table 8 is the source of truth for whether a code has an applicable
# BEGRF/ENDRF window (and which instrument sections were asked):
#   - FULL_ROUND: reference dates apply; use BEGRF→ENDRF (clamped to year).
#   - STOP: coverage ends mid-round at ENDRF.
#   - NO_COVERAGE: Table 8 beginning/ending dates are "Inapplicable".
#
# EXCLUDED_PSTATS: drop the person from the modeling cohort if ANY round
# carries one of these codes. Denominator math below still understands the
# full taxonomy so BEGRF/ENDRF month-year windows stay correct for statuses
# that remain (and for unit tests of death/stop paths).
EXCLUDED_PSTATS = {
    12, 13, 23, 24, 31, 32, 33, 35, 36, 43, 61, 62, 63, 64, 72, 73, 74, 81,
}
FULL_ROUND_STATUSES = {11, 12, 13, 14, 22, 41, 42, 44, 51, 71}
STOP_STATUSES = {32, 33, 34, 35, 36}
NO_COVERAGE_STATUSES = {0, 21, 24, 43, 62, 63, 64, 72, 73, 74, 81}


# ---------------------------------------------------------------------------
# Log data model
# ---------------------------------------------------------------------------

@dataclass
class StageEntry:
    stage: str
    rows_in: int
    rows_out: int
    patients_in: int | None = None
    patients_out: int | None = None
    detail: str = ""

    def to_dict(self) -> dict:
        return {
            "stage": self.stage,
            "rows_in": int(self.rows_in),
            "rows_out": int(self.rows_out),
            "rows_dropped": int(self.rows_in - self.rows_out),
            "patients_in": self.patients_in,
            "patients_out": self.patients_out,
            "patients_dropped": (
                None if self.patients_in is None or self.patients_out is None
                else int(self.patients_in - self.patients_out)
            ),
            "detail": self.detail,
        }


@dataclass
class RunLog:
    year: int
    meps_dir: str
    run_at: str = field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    stages: list[StageEntry] = field(default_factory=list)
    columns_dropped: dict[str, list[str]] = field(default_factory=dict)
    decisions: list[str] = field(default_factory=list)
    final_row_count: int | None = None
    final_patient_count: int | None = None

    def add(self, stage: StageEntry) -> None:
        self.stages.append(stage)

    def decide(self, note: str) -> None:
        self.decisions.append(note)

    def to_dict(self) -> dict:
        return {
            "year": self.year,
            "run_at": self.run_at,
            "meps_dir": self.meps_dir,
            "stages": [s.to_dict() for s in self.stages],
            "columns_dropped": self.columns_dropped,
            "decisions": self.decisions,
            "final_row_count": self.final_row_count,
            "final_patient_count": self.final_patient_count,
        }


# ---------------------------------------------------------------------------
# Path resolution
# ---------------------------------------------------------------------------

# App scripts now live at ``<repo>/app/`` (moved from ``Notebooks/MEPS/app/``).
# Outputs + decisions_log stay under ``Notebooks/MEPS/`` so historical data
# (parquets, tables, decisions_log) isn't orphaned. Data files stay under
# ``<repo>/data/MEPS/`` — resolved dynamically by ``resolve_meps_dir``.
APP_DIR = Path(__file__).resolve().parent
REPO_ROOT = APP_DIR.parent
NOTEBOOK_DIR = REPO_ROOT / "Notebooks" / "MEPS"
# Back-compat alias: pre-move code and ``write_log`` expect ``MEPS_DIR`` to
# point at the notebooks folder (where decisions_log.md lives). Keeping the
# name so importers (simple_app.py, tests) don't need to change.
MEPS_DIR = NOTEBOOK_DIR


def resolve_meps_dir(start: Path | None = None) -> Path:
    """Walk up from ``start`` (or CWD) looking for ``data/MEPS`` with year files."""
    cwd = (start or Path.cwd()).resolve()
    markers = [cfg["rx"] for cfg in YEAR_FILES.values()]
    candidates: list[Path] = []
    # CWD-relative candidates (works when invoked from any subdir)
    for base in [cwd, cwd.parent, cwd.parent.parent, cwd.parent.parent.parent]:
        candidates += [base / "data" / "MEPS" / "excels", base / "data" / "MEPS"]
    # Repo-relative fallback: repo_root/data/MEPS regardless of CWD
    candidates.append(REPO_ROOT / "data" / "MEPS" / "excels")
    candidates.append(REPO_ROOT / "data" / "MEPS")
    for d in candidates:
        if d.exists() and any((d / m).exists() for m in markers):
            return d
    raise FileNotFoundError(
        "Could not find MEPS data directory. Looked under: "
        + ", ".join(str(c) for c in candidates)
    )


def rx_prefix(year: int) -> str:
    return Path(YEAR_FILES[year]["rx"]).stem


def output_dirs(year: int) -> tuple[Path, Path]:
    base = NOTEBOOK_DIR / "output" / str(year)
    return base / "tables", base / "graphs"


def all_years_output_dirs() -> tuple[Path, Path]:
    base = NOTEBOOK_DIR / "output" / "all_years"
    return base / "tables", base / "graphs"


ALL_YEARS_PARQUET = "new_grouped_merge_df_chronic_drugs.parquet"
ALL_YEARS_XLSX = "new_grouped_merge_df_chronic_drugs.xlsx"
ALL_YEARS_FILTER_OPTIONS = "filter_options.json"
BRIDGE_PARQUET = "patient_drug_condition_bridge.parquet"
ALL_YEARS_BRIDGE_PARQUET = "patient_drug_condition_bridge.parquet"
MODEL_DF_ALL_YEARS = "model_df_all_years.parquet"

# Paper / Model-4 enrichments (sidebar demos + model medication features).
MAX_PILLS_PER_DAY = 10
PAPER_DEMO_COLS = ["MARRYXX", "REGIONXX", "EDUCYR", "RACETHX"]
PAPER_MED_COLS = [
    "medication_freq",
    "medication_dose",
    "medication_dose_unit",
    "medication_qty_unit",
    "medication_freq_bin",
]
PAPER_ENRICH_COLS = PAPER_DEMO_COLS + PAPER_MED_COLS


def _harmonize_year_columns(df: pd.DataFrame, year: int) -> pd.DataFrame:
    """Rename AGE/POVCAT/INSCOV{yy} → shared names and attach YEAR."""
    out = normalize_age_column(df, year).copy()
    yy = year % 100
    renames = {}
    for src, dest in (
        (f"AGE{yy:02d}X", "AGE"),
        (f"POVCAT{yy:02d}", "POVCAT"),
        (f"INSCOV{yy:02d}", "INSCOV"),
    ):
        if src in out.columns and dest not in out.columns:
            renames[src] = dest
    if renames:
        out = out.rename(columns=renames)
    out["YEAR"] = year
    return out


def _stringify_age_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Make AGE / AGEyyX parquet-safe (mixed int + ``unknown`` → string)."""
    out = df.copy()
    age_cols = [
        c for c in out.columns
        if c == "AGE" or (c.startswith("AGE") and c.endswith("X"))
    ]
    for col in age_cols:
        def _one(v):
            if pd.isna(v):
                return pd.NA
            if isinstance(v, str) and v.strip().lower() in {"unknown", "-1", "nan", "none"}:
                return "unknown"
            try:
                return str(int(float(v)))
            except (TypeError, ValueError):
                s = str(v).strip()
                return "unknown" if s.lower() in {"unknown", "-1"} else s

        out[col] = out[col].map(_one).astype("string")
    for col in ("medication_dose_unit", "medication_qty_unit", "medication_freq_bin"):
        if col in out.columns:
            out[col] = out[col].astype("string")
    return out


def _read_year_chronic_frame(year: int) -> pd.DataFrame | None:
    """Prefer per-year parquet cache; fall back to the Excel export.

    When only Excel exists, also write a parquet twin so the next merge is fast.
    """
    tables = output_dirs(year)[0]
    parquet = tables / "new_grouped_merge_df_chronic_drugs.parquet"
    xlsx = tables / "new_grouped_merge_df_chronic_drugs.xlsx"
    if parquet.exists():
        return pd.read_parquet(parquet)
    if xlsx.exists():
        try:
            df = pd.read_excel(xlsx, engine="calamine")
        except Exception:
            df = pd.read_excel(xlsx)
        try:
            _stringify_age_columns(df).to_parquet(parquet, index=False)
            print(f"[all_years]   cached year parquet -> {parquet.name}")
        except Exception as exc:
            print(f"[all_years]   warn: could not write year parquet: {exc}")
        return df
    return None


def _read_year_bridge_frame(year: int) -> pd.DataFrame | None:
    """Load the per-year patient_drug × condition bridge (parquet or xlsx)."""
    tables = output_dirs(year)[0]
    parquet = tables / BRIDGE_PARQUET
    xlsx = tables / "patient_drug_condition_bridge.xlsx"
    if parquet.exists():
        return pd.read_parquet(parquet)
    if xlsx.exists():
        try:
            df = pd.read_excel(xlsx, engine="calamine")
        except Exception:
            df = pd.read_excel(xlsx)
        try:
            df.to_parquet(parquet, index=False)
            print(f"[all_years]   cached bridge parquet -> {parquet.name}")
        except Exception as exc:
            print(f"[all_years]   warn: could not write bridge parquet: {exc}")
        return df
    return None


def _write_all_years_filter_options(merged: pd.DataFrame, tables_dir: Path) -> Path:
    """Sidebar filter metadata so the app need not scan the full frame."""
    import json

    age = pd.to_numeric(merged["AGE"], errors="coerce") if "AGE" in merged.columns else pd.Series(dtype=float)
    conditions = (
        sorted(merged["ICD10CDX_LABEL"].dropna().astype(str).unique().tolist())
        if "ICD10CDX_LABEL" in merged.columns
        else []
    )
    n_single = 0
    n_cond = len(conditions)
    if "ICD10CDX_LABEL" in merged.columns and "DUPERSID" in merged.columns:
        vc = merged.groupby("ICD10CDX_LABEL")["DUPERSID"].nunique()
        n_single = int((vc == 1).sum())
        n_cond = int(len(vc))

    def _valid_codes(col: str) -> list[int]:
        if col not in merged.columns:
            return []
        s = pd.to_numeric(merged[col], errors="coerce")
        return sorted(int(v) for v in s.dropna().unique() if v >= 0)

    educ = (
        pd.to_numeric(merged["EDUCYR"], errors="coerce")
        if "EDUCYR" in merged.columns
        else pd.Series(dtype=float)
    )
    educ_ok = educ[educ >= 0]

    payload = {
        "years": sorted(int(y) for y in merged["YEAR"].dropna().unique()) if "YEAR" in merged.columns else [],
        "n_rows": int(len(merged)),
        "n_patients": int(merged["DUPERSID"].nunique()) if "DUPERSID" in merged.columns else 0,
        "n_conditions": n_cond,
        "conditions_with_one_person": n_single,
        "age_min": int(age.min()) if len(age.dropna()) else 0,
        "age_max": int(age.max()) if len(age.dropna()) else 85,
        "conditions": conditions,
        "has_sex": "SEX" in merged.columns,
        "has_povcat": "POVCAT" in merged.columns,
        "has_inscov": "INSCOV" in merged.columns,
        "has_marryxx": "MARRYXX" in merged.columns,
        "has_regionxx": "REGIONXX" in merged.columns,
        "has_educyr": "EDUCYR" in merged.columns,
        "has_racethx": "RACETHX" in merged.columns,
        "marry_codes": _valid_codes("MARRYXX"),
        "region_codes": _valid_codes("REGIONXX"),
        "racethx_codes": _valid_codes("RACETHX"),
        "educyr_min": int(educ_ok.min()) if len(educ_ok) else 0,
        "educyr_max": int(educ_ok.max()) if len(educ_ok) else 17,
    }
    path = tables_dir / ALL_YEARS_FILTER_OPTIONS
    path.write_text(json.dumps(payload, indent=2))
    return path


def export_merged_all_years(meps_dir: str | Path | None = None) -> Path | None:
    """Build the all-years Streamlit cache under ``output/all_years/tables/``.

    Writes a fast **parquet** cache (what the app loads), plus filter-option JSON
    and a short summary. Run from the terminal::

        cd app
        python clean_meps.py cache-all-years

    Returns the parquet path, or ``None`` if no year exports exist yet.
    """
    _ = meps_dir  # reserved for API symmetry with run_exports
    tables_dir, _graphs_dir = all_years_output_dirs()
    tables_dir.mkdir(parents=True, exist_ok=True)

    frames: list[pd.DataFrame] = []
    bridge_frames: list[pd.DataFrame] = []
    for year in sorted(YEAR_FILES):
        print(f"[all_years] loading {year} …")
        df = _read_year_chronic_frame(year)
        if df is None:
            print(f"[all_years]   skip {year}: no chronic-drug export")
            continue
        frames.append(_harmonize_year_columns(df, year))
        print(f"[all_years]   {len(df):,} rows")

        br = _read_year_bridge_frame(year)
        if br is not None:
            br = br.copy()
            br["YEAR"] = year
            bridge_frames.append(br)
            print(f"[all_years]   bridge: {len(br):,} rows")
        else:
            print(f"[all_years]   no bridge yet for {year} — rerun `python "
                  f"clean_meps.py export --year {year}` after the CLNK-dedup fix")

    if not frames:
        print("[all_years] nothing to merge — run `python clean_meps.py export --year YYYY` first")
        return None

    merged = pd.concat(frames, ignore_index=True)
    merged = _stringify_age_columns(merged)

    parquet_path = tables_dir / ALL_YEARS_PARQUET
    merged.to_parquet(parquet_path, index=False)
    print(f"[all_years] wrote {parquet_path} ({len(merged):,} rows)")

    if bridge_frames:
        bridge_all = pd.concat(bridge_frames, ignore_index=True)
        bridge_path = tables_dir / ALL_YEARS_BRIDGE_PARQUET
        bridge_all.to_parquet(bridge_path, index=False)
        print(f"[all_years] wrote {bridge_path} ({len(bridge_all):,} rows)")

    # Optional Excel copy for manual inspection (slower; skip if huge preference)
    # Keep a slim summary xlsx instead of rewriting the full frame to Excel.

    if "ICD10CDX_LABEL" in merged.columns and "DUPERSID" in merged.columns:
        cond_counts = (
            merged.groupby("ICD10CDX_LABEL", as_index=False)
            .agg(n_patients=("DUPERSID", "nunique"))
            .sort_values("n_patients")
        )
        singles = cond_counts[cond_counts["n_patients"] == 1]
        summary = pd.DataFrame(
            {
                "metric": [
                    "years_included",
                    "rows",
                    "unique_patients",
                    "unique_conditions",
                    "conditions_with_one_person",
                ],
                "value": [
                    ",".join(str(y) for y in sorted(merged["YEAR"].unique())),
                    len(merged),
                    int(merged["DUPERSID"].nunique()),
                    int(len(cond_counts)),
                    int(len(singles)),
                ],
            }
        )
        summary.to_excel(tables_dir / "all_years_merge_summary.xlsx", index=False)
        singles.rename(columns={"ICD10CDX_LABEL": "condition name"}).to_excel(
            tables_dir / "conditions_one_person.xlsx", index=False
        )
        print(
            f"[all_years] conditions with 1 person: "
            f"{len(singles):,} / {len(cond_counts):,}"
        )

    opts = _write_all_years_filter_options(merged, tables_dir)
    print(f"[all_years] wrote {opts}")
    return parquet_path


# ---------------------------------------------------------------------------
# PSTATS reference-days computation
# ---------------------------------------------------------------------------

def _month_start(year, month):
    if pd.isna(month) or pd.isna(year) or month <= 0 or year <= 0:
        return None
    return date(int(year), int(month), 1)


def _month_end(year, month):
    if pd.isna(month) or pd.isna(year) or month <= 0 or year <= 0:
        return None
    return date(int(year), int(month), calendar.monthrange(int(year), int(month))[1])


def _union_days(intervals: list[tuple[date, date]]) -> tuple[int, date | None, date | None]:
    """Merge inclusive date intervals and return (n_days, overall_start, overall_end)."""
    if not intervals:
        return 0, None, None
    ordered = sorted((s, e) for s, e in intervals if e >= s)
    if not ordered:
        return 0, None, None
    merged: list[list[date]] = [[ordered[0][0], ordered[0][1]]]
    for start, end in ordered[1:]:
        # contiguous if start == last_end + 1 day
        if start <= merged[-1][1] + timedelta(days=1):
            if end > merged[-1][1]:
                merged[-1][1] = end
        else:
            merged.append([start, end])
    days = sum((e - s).days + 1 for s, e in merged)
    return days, merged[0][0], merged[-1][1]


def has_excluded_pstats(df: pd.DataFrame) -> pd.Series:
    """True when any of PSTATS31/42/53 is in ``EXCLUDED_PSTATS``."""
    mask = pd.Series(False, index=df.index)
    for sfx in (31, 42, 53):
        col = f"PSTATS{sfx}"
        if col not in df.columns:
            continue
        mask = mask | pd.to_numeric(df[col], errors="coerce").isin(EXCLUDED_PSTATS)
    return mask


def _compute_reference_coverage(row, year: int):
    """Return (eligible_days, coverage_start, coverage_end, notes).

    Eligible days are the union of each in-scope round's BEGRF/ENDRF
    month-year window, clamped to the calendar year. ``PSTATS == -1`` skips
    that round. Death / STOP statuses end coverage at that round's ENDRF.
    """
    year_start = date(year, 1, 1)
    year_end = date(year, 12, 31)

    ps_values = [int(row[f"PSTATS{sfx}"]) for sfx in (31, 42, 53)]
    if all(p == -1 for p in ps_values):
        return 0, None, None, "not_in_any_round"
    # Table 8: codes 11 and 12 share the same full-round reference-date rule.
    if all(p == 11 for p in ps_values):
        return 365, year_start, year_end, "full_year_all_rounds_11"
    if all(p == 12 for p in ps_values):
        return 365, year_start, year_end, "full_year_all_rounds_12"

    intervals: list[tuple[date, date]] = []
    notes: list[str] = []

    for sfx in (31, 42, 53):
        ps = int(row[f"PSTATS{sfx}"])
        beg = _month_start(row[f"BEGRFY{sfx}"], row[f"BEGRFM{sfx}"])
        end = _month_end(row[f"ENDRFY{sfx}"], row[f"ENDRFM{sfx}"])

        if ps == -1:
            notes.append(f"R{sfx}:not_in_round")
            continue
        if ps in NO_COVERAGE_STATUSES:
            notes.append(f"R{sfx}:no_coverage({ps})")
            continue

        start = max(beg, year_start) if beg else year_start

        if ps in FULL_ROUND_STATUSES:
            stop = min(end, year_end) if end else year_end
            if stop >= start:
                intervals.append((start, stop))
            notes.append(f"R{sfx}:active({ps})")
        elif ps == 31:
            stop = min(end, year_end) if end else year_end
            if stop >= start:
                intervals.append((start, stop))
            notes.append(f"R{sfx}:death")
            break
        elif ps in STOP_STATUSES:
            stop = min(end, year_end) if end else year_end
            if stop >= start:
                intervals.append((start, stop))
            notes.append(f"R{sfx}:stop({ps})")
            break
        else:
            notes.append(f"R{sfx}:UNCLASSIFIED({ps})")

    days, cs, ce = _union_days(intervals)
    if days == 0 or cs is None or ce is None:
        detail = ";".join(notes) if notes else "no_coverage"
        return 0, None, None, detail
    return days, cs, ce, ";".join(notes)


# ---------------------------------------------------------------------------
# Column sets (year-parameterized)
# ---------------------------------------------------------------------------

def _rx_cols(year: int) -> list[str]:
    yy = f"{year % 100:02d}"
    return [
        "DUPERSID", "DRUGIDX", "LINKIDX", "RXRECIDX", "RXDAYSUP",
        "RXNAME", "RXDRGNAM", "RXBEGYRX", "RXBEGMM", "RXNDC", "TC1", "TC1S1",
        "RXSTRENG", "RXSTRUNT", "RXQUANTY", "RXFORM",
        f"RXXP{yy}X", f"RXSF{yy}X",
    ]


def _backfill_newest_to_oldest(
    df: pd.DataFrame, cols_newest_to_oldest: list[str], out_name: str
) -> pd.DataFrame:
    """First non-negative value walking newest → oldest rounds."""
    present = [c for c in cols_newest_to_oldest if c in df.columns]
    if not present:
        df[out_name] = pd.NA
        return df
    series = [pd.to_numeric(df[c], errors="coerce") for c in present]
    out = series[0].copy()
    for s in series[1:]:
        need = out.isna() | (out < 0)
        out = out.where(~need, s)
    df[out_name] = out
    return df


def _clean_med_unit(s: pd.Series) -> pd.Series:
    out = s.astype(str).str.strip()
    bad = out.str.lower().isin({"-15", "nan", "none", "", "<na>"})
    return out.where(~bad)


def _medication_freq_bin(f) -> object:
    if pd.isna(f):
        return pd.NA
    if f <= 1.25:
        return "1x_daily"
    if f <= 2.25:
        return "2x_daily"
    if f <= 3.25:
        return "3x_daily"
    return "4x_plus"


def _first_non_null(s: pd.Series):
    s = s.dropna()
    return s.iloc[0] if len(s) else pd.NA


def _marry_region_round_cols(year: int) -> tuple[list[str], list[str]]:
    yy = f"{year % 100:02d}"
    marry = [f"MARRY{yy}X", "MARRY53X", "MARRY42X", "MARRY31X"]
    region = [f"REGION{yy}", f"REGION{yy}X", "REGION53", "REGION42", "REGION31"]
    return marry, region


def _attach_demo_enrichments(person: pd.DataFrame, year: int) -> pd.DataFrame:
    """Create MARRYXX / REGIONXX (backfilled) and coerce EDUCYR / RACETHX."""
    out = person.copy()
    marry_cols, region_cols = _marry_region_round_cols(year)
    out = _backfill_newest_to_oldest(out, marry_cols, "MARRYXX")
    out = _backfill_newest_to_oldest(out, region_cols, "REGIONXX")
    for c in PAPER_DEMO_COLS:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce")
        else:
            out[c] = pd.NA
    drop_rounds = [c for c in marry_cols + region_cols if c in out.columns]
    if drop_rounds:
        out = out.drop(columns=drop_rounds)
    return out


def _medication_features_from_fills(fills: pd.DataFrame) -> pd.DataFrame:
    """Person–drug medication_freq / dose / units from fill-level Rx fields."""
    needed = {"DUPERSID", "DRUGIDX", "RXSTRENG", "RXSTRUNT", "RXQUANTY", "RXFORM", "RXDAYSUP"}
    if not needed.issubset(fills.columns):
        return pd.DataFrame(
            columns=["DUPERSID", "DRUGIDX", *PAPER_MED_COLS]
        )

    dose = pd.to_numeric(fills["RXSTRENG"], errors="coerce")
    qty = pd.to_numeric(fills["RXQUANTY"], errors="coerce")
    days = pd.to_numeric(fills["RXDAYSUP"], errors="coerce")
    dose = dose.where(dose >= 0)
    qty_ok = qty.where(qty > 0)
    days_ok = days.where((days > 0) & (days < 990))
    freq = qty_ok / days_ok
    freq = freq.where((freq > 0) & (freq <= MAX_PILLS_PER_DAY))

    tmp = fills[["DUPERSID", "DRUGIDX"]].copy()
    tmp["medication_dose"] = dose
    tmp["medication_dose_unit"] = _clean_med_unit(fills["RXSTRUNT"])
    tmp["medication_freq"] = freq
    tmp["medication_qty_unit"] = _clean_med_unit(fills["RXFORM"])

    med = (
        tmp.groupby(["DUPERSID", "DRUGIDX"], as_index=False)
        .agg(
            medication_freq=("medication_freq", "mean"),
            medication_dose=("medication_dose", "mean"),
            medication_dose_unit=("medication_dose_unit", _first_non_null),
            medication_qty_unit=("medication_qty_unit", _first_non_null),
        )
    )
    med["medication_freq_bin"] = med["medication_freq"].map(_medication_freq_bin)
    return med


def _person_header(path: Path) -> list[str]:
    """First-row Excel header only (avoid full-sheet / nrows=0 calamine cost)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        row = next(wb.active.iter_rows(min_row=1, max_row=1, values_only=True))
    finally:
        wb.close()
    return [str(c) for c in row if c is not None]


def load_demo_extra_year(
    year: int, meps_dir: str | Path | None = None
) -> pd.DataFrame:
    """MARRYXX / REGIONXX (backfilled) + EDUCYR + RACETHX from that year's person file."""
    dir_ = Path(meps_dir) if meps_dir is not None else resolve_meps_dir()
    yy = f"{year % 100:02d}"
    person_path = dir_ / YEAR_FILES[year]["person"]
    header = _person_header(person_path)

    marry_pref = [f"MARRY{yy}X", "MARRY53X", "MARRY42X", "MARRY31X"]
    region_pref = [f"REGION{yy}", f"REGION{yy}X", "REGION53", "REGION42", "REGION31"]
    marry_cols = [c for c in marry_pref if c in header]
    region_cols = [c for c in region_pref if c in header]
    base = ["DUPERSID"]
    for c in ("EDUCYR", "RACETHX"):
        if c in header:
            base.append(c)
    usecols = list(dict.fromkeys(base + marry_cols + region_cols))

    demo = pd.read_excel(person_path, engine="calamine", usecols=usecols)
    demo = demo.drop_duplicates("DUPERSID").copy()
    demo = _backfill_newest_to_oldest(demo, marry_cols, "MARRYXX")
    demo = _backfill_newest_to_oldest(demo, region_cols, "REGIONXX")
    for c in PAPER_DEMO_COLS:
        if c in demo.columns:
            demo[c] = pd.to_numeric(demo[c], errors="coerce")
        else:
            demo[c] = pd.NA
    demo["YEAR"] = year
    return demo[["DUPERSID", "YEAR", *PAPER_DEMO_COLS]]


def load_med_extra_year(
    year: int, meps_dir: str | Path | None = None
) -> pd.DataFrame:
    """Person-level medication freq/dose/units from that year's Rx file (for model_df)."""
    dir_ = Path(meps_dir) if meps_dir is not None else resolve_meps_dir()
    rx_path = dir_ / YEAR_FILES[year]["rx"]
    rx_cols = [
        "DUPERSID", "DRUGIDX", "RXNAME",
        "RXSTRENG", "RXSTRUNT", "RXQUANTY", "RXFORM", "RXDAYSUP",
    ]
    rx = pd.read_excel(rx_path, engine="calamine", usecols=rx_cols)

    dose = pd.to_numeric(rx["RXSTRENG"], errors="coerce")
    qty = pd.to_numeric(rx["RXQUANTY"], errors="coerce")
    days = pd.to_numeric(rx["RXDAYSUP"], errors="coerce")
    dose = dose.where(dose >= 0)
    qty_ok = qty.where(qty > 0)
    days_ok = days.where((days > 0) & (days < 990))
    freq = qty_ok / days_ok
    freq = freq.where((freq > 0) & (freq <= MAX_PILLS_PER_DAY))

    rx = rx.copy()
    rx["medication_dose"] = dose
    rx["medication_dose_unit"] = _clean_med_unit(rx["RXSTRUNT"])
    rx["medication_freq"] = freq
    rx["medication_qty_unit"] = _clean_med_unit(rx["RXFORM"])

    med_by_drug = (
        rx.groupby(["DUPERSID", "RXNAME"], as_index=False)
        .agg(
            medication_freq=("medication_freq", "mean"),
            medication_dose=("medication_dose", "mean"),
            medication_dose_unit=("medication_dose_unit", _first_non_null),
            medication_qty_unit=("medication_qty_unit", _first_non_null),
        )
    )
    med_by_person = (
        med_by_drug.groupby("DUPERSID", as_index=False)
        .agg(
            medication_freq=("medication_freq", "mean"),
            medication_dose=("medication_dose", "mean"),
            medication_dose_unit=("medication_dose_unit", _first_non_null),
            medication_qty_unit=("medication_qty_unit", _first_non_null),
        )
    )
    med_by_person["YEAR"] = year
    return med_by_person


def _normalize_rxdrgname(rxdrgnam: pd.Series, rxname: pd.Series | None = None) -> pd.Series:
    """Clean MEPS ``RXDRGNAM`` into a modeling ``RXDRGNAME``.

    Sentinels (-1/-7/-8/-15 and blanks) are treated as missing. When a fallback
    ``RXNAME`` series is provided, missing values are filled from it so every
    fill still has a drug label for one-hot encoding.
    """
    raw = rxdrgnam.copy()
    num = pd.to_numeric(raw, errors="coerce")
    as_str = raw.astype(str).str.strip()
    bad = (
        num.notna() & (num < 0)
    ) | as_str.str.upper().isin(
        {"-1", "-7", "-8", "-15", "NAN", "NONE", "", "<NA>"}
    )
    out = as_str.where(~bad)
    out = out.replace({"nan": pd.NA, "None": pd.NA, "<NA>": pd.NA})
    if rxname is not None:
        fb = rxname.astype(str).str.strip()
        out = out.fillna(fb)
    return out.astype("string")


def _drug_start_days_in_year(
    year: int,
    first_year: pd.Series,
    first_month: pd.Series,
    survey_start_month: pd.Series | None = None,
) -> pd.Series:
    """Days from a drug's earliest fill start to Dec 31 of ``year``.

    - drug started before ``year``  -> 365 (drug active whole year)
    - drug started in ``year`` with a valid month -> (Dec 31 - first-of-month + 1)
    - drug started in ``year`` but month is missing/sentinel -> use
      ``survey_start_month`` (person's BEGRF / ``ref_start`` month) when
      available; otherwise 365
    - drug's earliest start-year is missing or > ``year`` -> 365 (conservative)
    """
    fy = pd.to_numeric(first_year, errors="coerce")
    fm = pd.to_numeric(first_month, errors="coerce")
    year_end = date(year, 12, 31)
    days_by_month = {
        m: (year_end - date(year, m, 1)).days + 1 for m in range(1, 13)
    }
    out = pd.Series(365, index=fy.index, dtype="int64")
    mask_med = fy.eq(year) & fm.between(1, 12)
    out.loc[mask_med] = fm.loc[mask_med].astype(int).map(days_by_month).astype("int64")
    if survey_start_month is not None:
        sm = pd.to_numeric(survey_start_month, errors="coerce")
        mask_survey = fy.eq(year) & ~fm.between(1, 12) & sm.between(1, 12)
        out.loc[mask_survey] = (
            sm.loc[mask_survey].astype(int).map(days_by_month).astype("int64")
        )
    return out


def _person_cols(year: int) -> list[str]:
    yy = f"{year % 100:02d}"
    marry_cols, _region_cols = _marry_region_round_cols(year)
    # REGION{yy}X is absent in 2020–2023 person files — use year-end REGION{yy}.
    region_usecols = [f"REGION{yy}", "REGION53", "REGION42", "REGION31"]
    return [
        "DUPERSID", "SEX", f"AGE{yy}X", f"INSCOV{yy}", f"POVCAT{yy}",
        f"FAMINC{yy}", "RACEV2X", "EDUCYR", "RACETHX",
        *marry_cols,
        *region_usecols,
        "PSTATS31", "PSTATS42", "PSTATS53",
        "BEGRFM31", "BEGRFY31", "BEGRFM42", "BEGRFY42", "BEGRFM53", "BEGRFY53",
        "ENDRFM31", "ENDRFY31", "ENDRFM42", "ENDRFY42", "ENDRFM53", "ENDRFY53",
    ]


def normalize_age_column(df: pd.DataFrame, year: int) -> pd.DataFrame:
    """Replace MEPS AGE sentinel -1 with the string ``unknown``."""
    col = f"AGE{year % 100:02d}X"
    if col not in df.columns:
        return df
    out = df.copy()
    age = out[col]
    # Keep numeric ages; map -1 (and already-string unknown) to "unknown"
    is_missing = age.astype(str).str.strip().isin({"-1", "unknown"}) | (age == -1)
    out[col] = age.where(~is_missing, other="unknown")
    return out


COND_COLS = ["DUPERSID", "CONDIDX", "ICD10CDX", "AGEDIAG"]


# ---------------------------------------------------------------------------
# Core pipeline stages (pure functions, unit-testable without MEPS files)
# ---------------------------------------------------------------------------

def link_rx_to_conditions(
    rx: pd.DataFrame,
    clnk: pd.DataFrame,
    h249_chronic: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """CLNK-link Rx fills to chronic conditions and produce the safe frames.

    Returns
    -------
    merged
        rx x chronic-condition rows, one row per (fill, linked chronic ICD).
        A fill linked to N chronic ICDs appears N times here. Use for
        condition-context aggregations only, never sum RXDAYSUP on this frame.
    fills
        Deduplicated on (DUPERSID, DRUGIDX, RXRECIDX) so each physical fill
        appears once. RXDAYSUP is safe to sum on this frame.
    bridge
        One row per (DUPERSID, DRUGIDX, ICD10CDX, CONDIDX). Carries no
        RXDAYSUP by design so it cannot be accidentally summed. Use for
        condition-level rollups.
    """
    cond_df = clnk.merge(
        h249_chronic.drop(columns=["DUPERSID"]),
        on="CONDIDX", how="left",
    )
    cond_df = cond_df[cond_df["EVENTYPE"] == 8]

    keep_cond_cols = ["DUPERSID", "CONDIDX", "EVNTIDX", "ICD10CDX",
                      "ICD10CDX_LABEL", "is_chronic"]
    merged = rx.merge(
        cond_df[keep_cond_cols],
        left_on=["DUPERSID", "LINKIDX"], right_on=["DUPERSID", "EVNTIDX"],
        how="left",
    )
    merged = merged.dropna(subset=["is_chronic"])

    fills = merged.drop_duplicates(subset=["DUPERSID", "DRUGIDX", "RXRECIDX"])

    bridge = (
        merged.dropna(subset=["ICD10CDX"])[
            ["DUPERSID", "DRUGIDX", "ICD10CDX", "ICD10CDX_LABEL",
             "is_chronic", "CONDIDX"]
        ]
        .drop_duplicates(subset=["DUPERSID", "DRUGIDX", "ICD10CDX", "CONDIDX"])
        .reset_index(drop=True)
    )
    return merged, fills, bridge


def compute_reference_days(person: pd.DataFrame, year: int) -> pd.DataFrame:
    """PSTATS + BEGRF/ENDRF -> per-person eligible days in ``year``.

    Wraps ``_compute_reference_coverage`` over one person-year frame. Returns
    one row per DUPERSID with:
    ``total_days_supply``, ``ref_start_{year}``, ``ref_end_{year}``,
    ``participation_type``, ``coverage_notes``, ``r53_nonresponse``.
    """
    person_demo = person.drop_duplicates("DUPERSID").copy()
    parts = person_demo.apply(
        lambda row: _compute_reference_coverage(row, year),
        axis=1, result_type="expand",
    )
    person_demo["total_days_supply"] = parts[0]
    person_demo[f"ref_start_{year}"] = parts[1]
    person_demo[f"ref_end_{year}"] = parts[2]
    person_demo["coverage_notes"] = parts[3]
    person_demo["r53_nonresponse"] = person_demo["PSTATS53"] == -1

    person_demo["participation_type"] = np.select(
        [
            person_demo["coverage_notes"].str.startswith(
                "full_year_all_rounds", na=False
            ),
            person_demo["coverage_notes"].eq("not_in_any_round"),
            person_demo["coverage_notes"].str.contains("death", na=False),
            person_demo["coverage_notes"].str.contains("stop", na=False),
            person_demo["coverage_notes"].str.contains("no_coverage", na=False),
            person_demo["total_days_supply"].eq(0),
        ],
        ["full_year", "not_in_any_round", "ended_early_death",
         "ended_early_left_ru", "partial_no_survey", "no_coverage"],
        default="partial_year",
    )

    return person_demo[
        ["DUPERSID", f"ref_start_{year}", f"ref_end_{year}",
         "total_days_supply", "participation_type", "coverage_notes",
         "r53_nonresponse"]
    ]


def apply_adherence_math(
    df: pd.DataFrame,
    denom_col: str = "total_days_supply",
) -> pd.DataFrame:
    """Add ``total_valid_days`` and ``meps_adherence_ratio`` in place-safe way.

    Formula (PDC-style, 100% ceiling):
        total_valid_days = min(RXDAYSUP, 365, denom_col)
        ratio            = 100 * total_valid_days / denom_col  (NaN if denom == 0)

    Requires ``RXDAYSUP`` and ``denom_col`` on ``df``. Returns the same frame
    with the two columns added/overwritten.
    """
    out = df.copy()
    out["total_valid_days"] = np.minimum(out["RXDAYSUP"], 365).astype(int)
    out["total_valid_days"] = np.minimum(out["total_valid_days"], out[denom_col])
    out["meps_adherence_ratio"] = np.where(
        out[denom_col].eq(0),
        np.nan,
        out["total_valid_days"] / out[denom_col] * 100,
    )
    return out


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build(
    year: int,
    meps_dir: str | Path | None = None,
    drug_chronic_only: bool = True,
    pstats_denominator: bool = True,
) -> tuple[pd.DataFrame, pd.DataFrame, RunLog]:
    """Build ``new_grouped_merge_df`` for one year.

    Parameters
    ----------
    year
        One of 2020, 2021, 2022, 2023.
    meps_dir
        Directory with the MEPS excels. Auto-detected by ``resolve_meps_dir``
        if not passed.
    drug_chronic_only
        If True (default), filter to chronic drugs per Friana's labeled CSV.
        If False, keep all drugs and let the caller filter.
    pstats_denominator
        If True (default), use PSTATS-based reference days from the person file.
        If False, use a flat 365-day denominator (notebook cells 11-15 style).

    Returns
    -------
    (patient_drug, bridge, log)
        ``patient_drug`` has one row per (DUPERSID, DRUGIDX) with adherence
        and reference-days columns; ``primary_ICD10CDX`` and its alias
        ``ICD10CDX`` carry the first-encountered chronic ICD, and
        ``n_chronic_conditions`` + ``chronic_conditions`` (comma-joined)
        preserve the full linkage set. ``bridge`` is one row per (DUPERSID,
        DRUGIDX, ICD10CDX, CONDIDX) — use it for condition-level rollups.
        ``log`` records every stage's counts and decisions — feed it to
        ``write_log(log)`` to append to ``decisions_log.md``.
    """
    if year not in YEAR_FILES:
        raise ValueError(f"year must be one of {sorted(YEAR_FILES)}; got {year}")

    dir_ = Path(meps_dir) if meps_dir is not None else resolve_meps_dir()
    log = RunLog(year=year, meps_dir=str(dir_))
    files = YEAR_FILES[year]
    yy = f"{year % 100:02d}"
    xp_col = f"RXXP{yy}X"
    sf_col = f"RXSF{yy}X"

    log.decide(
        f"Denominator is PSTATS-derived reference days from {files['person']} "
        f"(BEGRF/ENDRF per round). Full-year respondents get 365; deceased "
        f"and moved-out get windowed."
    )
    log.decide(
        "PSTATS taxonomy covers 25 documented codes per h251doc Tables 7-8; "
        "unknown codes marked 'UNCLASSIFIED' rather than silently included."
    )

    # -- 1. Load h248a, keep only lean columns --------------------------
    h248a = pd.read_excel(dir_ / files["rx"], engine="calamine",
                          usecols=_rx_cols(year))
    log.add(StageEntry("load_rx", rows_in=len(h248a), rows_out=len(h248a),
                       patients_in=h248a["DUPERSID"].nunique(),
                       patients_out=h248a["DUPERSID"].nunique(),
                       detail=f"file={files['rx']}"))

    # -- 2. RXDAYSUP filter: keep 1..989 (drops 999 'as needed' + negs) --
    before, before_p = len(h248a), h248a["DUPERSID"].nunique()
    rx = h248a[(h248a["RXDAYSUP"] > 0) & (h248a["RXDAYSUP"] < 990)]
    log.add(StageEntry("filter_RXDAYSUP_1_to_989",
                       rows_in=before, rows_out=len(rx),
                       patients_in=before_p, patients_out=rx["DUPERSID"].nunique(),
                       detail="Excludes RXDAYSUP == 999 (MEPS 'as needed' flag) and "
                              "negative sentinels (-1, -7, -8, -9, -15)."))
    log.decide(
        "RXDAYSUP == 999 rows dropped: 999 is a flag for 'as-needed' meds, not "
        "a day count. Treating it as adherence data would create fake 100% MPR."
    )

    # -- 3. Drop rows with unusable start-year info ---------------------
    before, before_p = len(rx), rx["DUPERSID"].nunique()
    rx = rx[rx["RXBEGYRX"] > 0]
    log.add(StageEntry("filter_RXBEGYRX_positive",
                       rows_in=before, rows_out=len(rx),
                       patients_in=before_p, patients_out=rx["DUPERSID"].nunique(),
                       detail="Excludes rows where the start-year variable is a "
                              "MEPS sentinel (-1, -7, -8)."))
    log.decide(
        "Fills with unknown RXBEGYRX (~28% of raw rows) are excluded because "
        "the denominator branch can't route them. Loses coverage for meds "
        "with missing history; documented as a limitation."
    )

    # -- 4. Load h249 conditions + is_chronic ICD list ------------------
    h249 = pd.read_excel(dir_ / files["cond"], engine="calamine",
                         usecols=COND_COLS)
    is_chronic_icd = pd.read_excel(dir_ / CHRONIC_ICD_FILE, engine="calamine")
    log.decide(
        f"Chronic-condition filter uses {CHRONIC_ICD_FILE} (Friana's DRAFT "
        "ICD-10 3-digit allowlist). The Data Reference Guide flags this file "
        "as needing verification — treat downstream ICD-level counts with "
        "care until it's audited."
    )
    h249_ic = h249.merge(
        is_chronic_icd[["ICD10CDX", "ICD10CDX_LABEL", "is_chronic"]],
        on="ICD10CDX", how="left",
    )
    log.add(StageEntry("merge_chronic_icd_list",
                       rows_in=len(h249), rows_out=len(h249_ic),
                       detail="Added is_chronic + ICD10CDX_LABEL to h249 rows."))

    before, before_p = len(h249_ic), h249_ic["DUPERSID"].nunique()
    h249_chronic = h249_ic[h249_ic["is_chronic"] > 0].copy()
    log.add(StageEntry("filter_conditions_to_chronic",
                       rows_in=before, rows_out=len(h249_chronic),
                       patients_in=before_p,
                       patients_out=h249_chronic["DUPERSID"].nunique(),
                       detail="Keeps only conditions whose ICD-10 3-digit code "
                              "is in the chronic allowlist."))

    # -- 5-6-6b-7b. CLNK link + dedup fills + bridge (extracted) --------
    # See ``link_rx_to_conditions`` for the join / dedup / bridge invariants.
    clnk = pd.read_excel(dir_ / files["clnk"], engine="calamine")
    rx_in, rx_p = len(rx), rx["DUPERSID"].nunique()
    merged, fills, bridge = link_rx_to_conditions(rx, clnk, h249_chronic)
    log.add(StageEntry("link_rx_to_chronic_conditions",
                       rows_in=rx_in, rows_out=len(merged),
                       patients_in=rx_p,
                       patients_out=merged["DUPERSID"].nunique(),
                       detail="CLNK EVENTYPE=8, LEFT JOIN rx on "
                              "(DUPERSID, LINKIDX=EVNTIDX), drop non-chronic. "
                              "Multi-condition fills expand to N rows here; "
                              "unwound in the fills dedup below."))
    log.add(StageEntry("dedup_fills_before_summing_days",
                       rows_in=len(merged), rows_out=len(fills),
                       patients_in=merged["DUPERSID"].nunique(),
                       patients_out=fills["DUPERSID"].nunique(),
                       detail="One row per (DUPERSID, DRUGIDX, RXRECIDX). "
                              "Removes CLNK multi-condition duplication so "
                              "RXDAYSUP is not double-counted when a fill "
                              "links to more than one chronic ICD."))
    log.decide(
        "Rx→condition join uses (DUPERSID, LINKIDX=EVNTIDX) — never DUPERSID "
        "alone (Cartesian explosion) and never LINKIDX alone (round-scoped). "
        "Documented in MEPS_SCHEMA_NOTES."
    )
    log.decide(
        "CLNK multi-condition duplication fix: RXDAYSUP is summed on "
        "deduplicated fills (one row per RXRECIDX per person-drug), not on "
        "the rx×condition join. Previous behaviour inflated days when a fill "
        "linked to multiple chronic conditions and silently kept only the "
        "first ICD. Multi-condition context is preserved via "
        "n_chronic_conditions + chronic_conditions list on patient_drug and "
        "the patient_drug_condition bridge frame."
    )

    # -- 7. Groupby to person-drug grain -------------------------------
    # Valid-month column for drug-start-date denominator: keep 1..12, mask sentinels
    fills = fills.assign(
        _valid_month=fills["RXBEGMM"].where(
            (fills["RXBEGMM"] >= 1) & (fills["RXBEGMM"] <= 12)
        )
    )
    gm = fills.groupby(["DUPERSID", "DRUGIDX"]).agg(
        RXDAYSUP=("RXDAYSUP", "sum"),
        RXXP=(xp_col, "mean"),
        RXSF=(sf_col, "mean"),
        RXNAME=("RXNAME", "first"),
        RXDRGNAM=("RXDRGNAM", "first"),
        RXNDC=("RXNDC", "first"),
        RXBEGYRX=("RXBEGYRX", "min"),
        first_month=("_valid_month", "min"),
        TC1=("TC1", "first"),
        TC1S1=("TC1S1", "first"),
        primary_ICD10CDX=("ICD10CDX", "first"),
        primary_ICD10CDX_LABEL=("ICD10CDX_LABEL", "first"),
    ).reset_index()
    gm = gm.rename(columns={"RXXP": xp_col, "RXSF": sf_col})
    gm["RXDRGNAME"] = _normalize_rxdrgname(gm["RXDRGNAM"], gm["RXNAME"])
    gm = gm.drop(columns=["RXDRGNAM"])
    # Preliminary drug-start days (no survey-start fallback yet — that needs
    # ref_start from PSTATS/BEGRF, applied after the reference-days merge).
    gm["drug_start_days"] = _drug_start_days_in_year(
        year, gm["RXBEGYRX"], gm["first_month"]
    )

    # Attach multi-condition context from the un-deduped merged frame so we
    # keep every chronic ICD the fill was CLNK-linked to.
    cond_context = (
        merged.dropna(subset=["ICD10CDX"])
        .groupby(["DUPERSID", "DRUGIDX"])
        .agg(
            n_chronic_conditions=("ICD10CDX", "nunique"),
            chronic_conditions=(
                "ICD10CDX",
                lambda s: ",".join(sorted(s.astype(str).unique())),
            ),
        )
        .reset_index()
    )
    gm = gm.merge(cond_context, on=["DUPERSID", "DRUGIDX"], how="left")
    gm["n_chronic_conditions"] = gm["n_chronic_conditions"].fillna(0).astype(int)
    gm["chronic_conditions"] = gm["chronic_conditions"].fillna("")

    # Backwards-compat aliases so notebooks, run_exports and the all-years
    # cache keep working. primary_* names make it explicit that these are
    # a choice, not the full picture.
    gm["ICD10CDX"] = gm["primary_ICD10CDX"]
    gm["ICD10CDX_LABEL"] = gm["primary_ICD10CDX_LABEL"]

    # Medication pattern features (person–drug grain) for Model-4 / paper enrichments
    med = _medication_features_from_fills(fills)
    if not med.empty:
        gm = gm.drop(columns=[c for c in PAPER_MED_COLS if c in gm.columns], errors="ignore")
        gm = gm.merge(med, on=["DUPERSID", "DRUGIDX"], how="left")

    log.add(StageEntry("groupby_person_drug",
                       rows_in=len(fills), rows_out=len(gm),
                       patients_in=fills["DUPERSID"].nunique(),
                       patients_out=gm["DUPERSID"].nunique(),
                       detail="One row per (DUPERSID, DRUGIDX). RXDAYSUP is "
                              "summed on the deduped fills; ICD10CDX = "
                              "primary_ICD10CDX kept as first-encountered "
                              "chronic ICD (aliased so downstream code that "
                              "reads ICD10CDX keeps working); "
                              "n_chronic_conditions + chronic_conditions list "
                              "cover the full linkage set; medication_freq / "
                              "dose / units attached from RXQUANTY/RXDAYSUP / "
                              "RXSTRENG / RXFORM."))

    # -- 7b. Patient-drug × condition bridge (built by link_rx_to_conditions)
    log.add(StageEntry("build_patient_drug_condition_bridge",
                       rows_in=len(merged), rows_out=len(bridge),
                       patients_in=merged["DUPERSID"].nunique(),
                       patients_out=bridge["DUPERSID"].nunique(),
                       detail="One row per (DUPERSID, DRUGIDX, ICD10CDX, "
                              "CONDIDX). Bridge for condition-level rollups; "
                              "carries no RXDAYSUP so cannot be summed by "
                              "accident."))

    if not pstats_denominator:
        gm["total_days_supply"] = np.where(
            gm["RXBEGYRX"] <= year, gm["drug_start_days"], 0
        ).astype(int)
        gm = apply_adherence_math(gm, denom_col="total_days_supply")
        log.decide(
            "Flat-365 denominator with drug-start adjust: total_days_supply = "
            "drug_start_days (365 if drug started before this year; else days "
            "from first-fill month to Dec 31) when RXBEGYRX <= year."
        )
        log.final_row_count = int(len(gm))
        log.final_patient_count = int(gm["DUPERSID"].nunique())
        bridge = bridge.merge(
            gm[["DUPERSID", "DRUGIDX"]].drop_duplicates(),
            on=["DUPERSID", "DRUGIDX"], how="inner",
        )
        return gm, bridge, log

    # -- 8. Optional: drug-chronic filter ------------------------------
    if drug_chronic_only:
        rxlabel = pd.read_csv(dir_ / CHRONIC_DRUG_FILE)
        rx_flag = rxlabel[["RXNAME", "is_chronic"]].rename(
            columns={"is_chronic": "is_chronic_drug"}
        )
        before, before_p = len(gm), gm["DUPERSID"].nunique()
        gm = gm.merge(rx_flag, on="RXNAME", how="left")
        gm = gm[gm["is_chronic_drug"] == 1].drop(columns=["is_chronic_drug"])
        log.add(StageEntry("filter_drug_chronic_only",
                           rows_in=before, rows_out=len(gm),
                           patients_in=before_p,
                           patients_out=gm["DUPERSID"].nunique(),
                           detail=f"Uses {CHRONIC_DRUG_FILE}. Drops drugs "
                                  "labeled 'flare_up' or 'non_chronic'. "
                                  "Keeps 'chronic'."))
        log.decide(
            "Drug-side chronic filter uses Friana's hand-labeled RXNAME list "
            "(1,158 unique drug names, 518 chronic, 373 flare_up, 267 "
            "non_chronic). Prevents acute meds (antibiotics, steroid bursts, "
            "topical FLUOROURACIL) from contaminating chronic-condition MPR."
        )

    # -- 9. Load h251, keep lean cols, merge on DUPERSID ---------------
    person = pd.read_excel(dir_ / files["person"], engine="calamine",
                           usecols=_person_cols(year))
    person = _attach_demo_enrichments(person, year)
    log.decide(
        f"Person-level columns pulled from {files['person']}: AGE{yy}X, "
        f"SEX, RACEV2X, RACETHX, EDUCYR, MARRYXX/REGIONXX (newest→oldest "
        f"round backfill), INSCOV{yy}, POVCAT{yy}, FAMINC{yy}, plus "
        "PSTATS/BEGRF/ENDRF for reference-days computation. Medication "
        "freq/dose/units come from the Rx file at person–drug grain."
    )

    # Drop persons with excluded disposition codes in any round before
    # denom/adherence math (death, institutionalization, FT military 12,
    # student-away 13, ineligible, etc.).
    person = person.drop_duplicates("DUPERSID").copy()
    excl_mask = has_excluded_pstats(person)
    n_excl_persons = int(excl_mask.sum())
    excluded_ids = set(person.loc[excl_mask, "DUPERSID"].astype(str))
    before_p_excl, before_gm_excl = person["DUPERSID"].nunique(), len(gm)
    person = person.loc[~excl_mask].copy()
    gm = gm[~gm["DUPERSID"].astype(str).isin(excluded_ids)].copy()
    log.add(StageEntry(
        "drop_excluded_pstats",
        rows_in=before_gm_excl,
        rows_out=len(gm),
        patients_in=before_p_excl,
        patients_out=gm["DUPERSID"].nunique(),
        detail=(
            f"Dropped persons with any PSTATS31/42/53 in {sorted(EXCLUDED_PSTATS)} "
            f"({n_excl_persons} person-file rows; "
            f"{before_gm_excl - len(gm)} person-drug rows removed)."
        ),
    ))
    log.decide(
        "Excluded PSTATS (any round): "
        f"{sorted(EXCLUDED_PSTATS)}. These persons are removed from the "
        "chronic-drug modeling cohort before BEGRF/ENDRF denominator math."
    )

    before, before_p = len(gm), gm["DUPERSID"].nunique()
    gm = gm.merge(person, on="DUPERSID", how="left", validate="many_to_one")
    gm = normalize_age_column(gm, year)
    log.decide(
        f"AGE{yy}X sentinel -1 replaced with string 'unknown' (MEPS missing/"
        "inapplicable age)."
    )
    log.add(StageEntry("merge_person_demographics",
                       rows_in=before, rows_out=len(gm),
                       patients_in=before_p,
                       patients_out=gm["DUPERSID"].nunique(),
                       detail=f"LEFT JOIN person {files['person']} on DUPERSID; "
                              "validated many_to_one to catch dupe person rows; "
                              f"AGE{yy}X -1 → 'unknown'."))

    # -- 10. PSTATS → reference_days_df (extracted) --------------------
    reference_days_df = compute_reference_days(person, year)
    log.add(StageEntry("compute_reference_days",
                       rows_in=len(person),
                       rows_out=len(reference_days_df),
                       patients_in=person["DUPERSID"].nunique(),
                       patients_out=reference_days_df["DUPERSID"].nunique(),
                       detail="PSTATS + BEGRF/ENDRF → per-person eligible "
                              f"days in {year}. Clamped to [Jan 1, Dec 31]."))
    n_r53 = int(reference_days_df["r53_nonresponse"].sum())
    log.decide(
        f"R5/3 nonresponders flagged (PSTATS53 == -1): {n_r53} persons. Their "
        "denominator ends at R4/2 interview date but their numerator can "
        "include RXDAYSUP that extends past it — MPR biased HIGH for this "
        "group. Kept in the frame; filter downstream if the analysis needs to."
    )

    # -- 11. Merge reference_days_df + compute adherence ----------------
    gm = gm.merge(reference_days_df, on="DUPERSID", how="left")

    # Recompute drug_start_days with survey-start fallback now that ref_start
    # is available. Same-year fills with missing RXBEGMM use the person's
    # BEGRF / ref_start month (matches 2023_clean intent).
    ref_col = f"ref_start_{year}"
    survey_month = (
        pd.to_datetime(gm[ref_col], errors="coerce").dt.month
        if ref_col in gm.columns
        else None
    )
    gm["drug_start_days"] = _drug_start_days_in_year(
        year,
        gm["RXBEGYRX"],
        gm["first_month"],
        survey_start_month=survey_month,
    )

    # Shrink the person-year window to the drug's own eligibility window.
    # Prior-year starts keep drug_start_days=365 so PSTATS days bind.
    gm["total_days_supply"] = np.minimum(
        gm["total_days_supply"], gm["drug_start_days"]
    ).astype(int)

    gm = apply_adherence_math(gm, denom_col="total_days_supply")
    log.decide(
        "meps_adherence_ratio caps numerator at min(sum RXDAYSUP, 365, "
        "eligible_days). This is PDC-style (100% ceiling), not MPR-style "
        "(uncapped). Documented; can be relaxed later if uncapped MPR is "
        "wanted."
    )
    log.decide(
        "Drug-start denominator adjust: RXBEGYRX < year → 365. "
        "RXBEGYRX == year with RXBEGMM in 1..12 → days from that month to "
        "Dec 31. RXBEGYRX == year with missing/sentinel month → use the "
        "person's survey start month (ref_start / BEGRF). Final "
        "total_days_supply = min(PSTATS eligible days, drug_start_days)."
    )

    # Drop round-detail cols the caller doesn't need
    marry_rounds, region_rounds = _marry_region_round_cols(year)
    drop_cols = [c for c in [
        "PSTATS31", "PSTATS42", "PSTATS53",
        "BEGRFM31", "BEGRFY31", "ENDRFM31", "ENDRFY31",
        "BEGRFM42", "BEGRFY42", "ENDRFM42", "ENDRFY42",
        "BEGRFM53", "BEGRFY53", "ENDRFM53", "ENDRFY53",
        *marry_rounds,
        *region_rounds,
    ] if c in gm.columns]
    if drop_cols:
        gm = gm.drop(columns=drop_cols)
        log.columns_dropped["person_year_round_detail"] = drop_cols

    log.final_row_count = int(len(gm))
    log.final_patient_count = int(gm["DUPERSID"].nunique())

    bridge = bridge.merge(
        gm[["DUPERSID", "DRUGIDX"]].drop_duplicates(),
        on=["DUPERSID", "DRUGIDX"], how="inner",
    )
    return gm, bridge, log


# ---------------------------------------------------------------------------
# Log persistence
# ---------------------------------------------------------------------------

def write_log(log: RunLog | dict, path: str | Path | None = None) -> Path:
    """Append a human-readable run block to ``decisions_log.md``.

    ``path`` defaults to ``<this file's dir>/decisions_log.md``.
    """
    d = log if isinstance(log, dict) else log.to_dict()
    if path is None:
        path = MEPS_DIR / "decisions_log.md"
    path = Path(path)

    existing = path.read_text() if path.exists() else ""
    has_our_header = "# Decisions Log" in existing

    md_lines: list[str] = []
    if not existing:
        # Brand-new file: write full header.
        md_lines += [
            "# Decisions Log",
            "",
            "Auto-appended by `clean_meps.build()`. Each block records one "
            "invocation of the pipeline — rows and patients lost at every "
            "stage, columns dropped, and decisions taken.",
            "",
        ]
    elif not has_our_header:
        # File exists with unrelated content (e.g. hand-written notes).
        # Preserve it; add a divider + section header before the first run block.
        md_lines += [
            "",
            "---",
            "",
            "# Decisions Log",
            "",
            "Auto-appended by `clean_meps.build()`. Each block records one "
            "invocation of the pipeline — rows and patients lost at every "
            "stage, columns dropped, and decisions taken.",
            "",
        ]

    md_lines += [
        f"## {d['year']} build — {d['run_at']}",
        "",
        f"**Source dir**: `{d['meps_dir']}`",
        f"**Final rows**: {d['final_row_count']:,}    "
        f"**Final unique patients**: {d['final_patient_count']:,}",
        "",
        "### Pipeline stages",
        "",
        "| Stage | Rows in | Rows out | Rows dropped | Patients in | Patients out | Patients dropped |",
        "| --- | --- | --- | --- | --- | --- | --- |",
    ]
    for s in d["stages"]:
        pi = s["patients_in"] if s["patients_in"] is not None else "—"
        po = s["patients_out"] if s["patients_out"] is not None else "—"
        pd_ = s["patients_dropped"] if s["patients_dropped"] is not None else "—"
        pi_s = f"{pi:,}" if isinstance(pi, int) else pi
        po_s = f"{po:,}" if isinstance(po, int) else po
        pd_s = f"{pd_:,}" if isinstance(pd_, int) else pd_
        md_lines.append(
            f"| `{s['stage']}` | {s['rows_in']:,} | {s['rows_out']:,} | "
            f"{s['rows_dropped']:,} | {pi_s} | {po_s} | {pd_s} |"
        )

    md_lines += ["", "### Detail per stage", ""]
    for s in d["stages"]:
        if s.get("detail"):
            md_lines.append(f"- **{s['stage']}**: {s['detail']}")

    if d["columns_dropped"]:
        md_lines += ["", "### Columns dropped", ""]
        for group, cols in d["columns_dropped"].items():
            md_lines.append(f"- **{group}**: `{', '.join(cols)}`")

    md_lines += ["", "### Decisions taken", ""]
    for note in d["decisions"]:
        md_lines.append(f"- {note}")

    md_lines += ["", "---", ""]

    # Always append — never overwrite. Header logic above decides whether to
    # prepend divider + header lines before the run block.
    with path.open("a") as f:
        f.write("\n".join(md_lines) + "\n")

    return path


# ---------------------------------------------------------------------------
# Notebook export runner (tables + graphs -> output/<year>/)
# ---------------------------------------------------------------------------

TOTAL_EXPORT_STEPS = 12


@dataclass
class ExportCtx:
    year: int
    tables_dir: Path
    graphs_dir: Path
    rx_p: str
    step_num: int = 0

    def log(self, msg: str) -> None:
        print(msg, flush=True)

    def step(self, title: str) -> None:
        self.step_num += 1
        self.log(f"\n[{self.step_num}/{TOTAL_EXPORT_STEPS}] {title}")

    def info(self, msg: str) -> None:
        self.log(f"    {msg}")


def _save_table(ctx: ExportCtx, df: pd.DataFrame, name: str) -> None:
    ctx.tables_dir.mkdir(parents=True, exist_ok=True)
    path = ctx.tables_dir / name
    if path.suffix == ".csv":
        df.to_csv(path, index=False)
    else:
        df.to_excel(path, index=False)
    ctx.info(f"saved table -> output/{ctx.year}/tables/{name}  ({len(df):,} rows)")


def _save_figure(ctx: ExportCtx, fig, name: str) -> None:
    ctx.graphs_dir.mkdir(parents=True, exist_ok=True)
    path = ctx.graphs_dir / name
    fig.savefig(path, dpi=150, bbox_inches="tight")
    import matplotlib.pyplot as plt

    plt.close(fig)
    ctx.info(f"saved graph -> output/{ctx.year}/graphs/{name}")


def _plot_adherence_histogram(
    ctx: ExportCtx,
    values: pd.Series,
    *,
    title: str,
    xlabel: str,
    ylabel: str,
    filename: str,
    adherence_threshold: int = 60,
) -> None:
    import matplotlib.pyplot as plt

    bins = np.arange(0, 110, 10)
    fig, ax = plt.subplots(figsize=(12, 6))
    _, edges, patches = ax.hist(values, bins=bins, edgecolor="black", linewidth=0.6)
    for patch, left, right in zip(patches, edges[:-1], edges[1:]):
        if right <= adherence_threshold:
            patch.set_facecolor("#d9534f")
        elif left >= adherence_threshold:
            patch.set_facecolor("#5cb85c")
        else:
            patch.set_facecolor("#f0ad4e")
    ax.axvline(
        adherence_threshold,
        color="red",
        linestyle=":",
        linewidth=2,
        label=f"{adherence_threshold}% adherence threshold",
    )
    bin_labels = [f"{int(left)}-{int(right)}" for left, right in zip(edges[:-1], edges[1:])]
    ax.set_xticks(edges[:-1] + 5)
    ax.set_xticklabels(bin_labels, rotation=45, ha="right")
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)
    ax.legend()
    plt.tight_layout()
    _save_figure(ctx, fig, filename)


def _condition_view(
    patient_drug: pd.DataFrame, bridge: pd.DataFrame
) -> pd.DataFrame:
    """Bridge-join view for condition-level rollups.

    One row per (person-drug × linked chronic condition). Adherence is carried
    from patient_drug; groupby ICD10CDX on the result gives the correct mean.
    """
    drop_from_pd = [c for c in [
        "ICD10CDX", "ICD10CDX_LABEL",
        "primary_ICD10CDX", "primary_ICD10CDX_LABEL",
    ] if c in patient_drug.columns]
    return bridge.merge(
        patient_drug.drop(columns=drop_from_pd),
        on=["DUPERSID", "DRUGIDX"], how="inner",
    )


def _export_summaries(
    ctx: ExportCtx,
    df: pd.DataFrame,
    bridge: pd.DataFrame,
    *,
    tag: str,
    title_note: str = "",
    adherence_threshold: int = 60,
    low_threshold: int = 10,
) -> None:
    note = f" ({title_note})" if title_note else ""
    cond_view = _condition_view(df, bridge)
    by_condition = (
        cond_view.groupby(["ICD10CDX", "ICD10CDX_LABEL"], as_index=False)
        .agg(meps_adherence_ratio=("meps_adherence_ratio", "mean"))
    )
    by_tc1s1 = (
        df.groupby(["TC1", "TC1S1"], as_index=False)
        .agg(meps_adherence_ratio=("meps_adherence_ratio", "mean"))
    )
    n_adherent = (by_condition["meps_adherence_ratio"] >= adherence_threshold).sum()
    _plot_adherence_histogram(
        ctx,
        by_condition["meps_adherence_ratio"],
        title=(
            f"Distribution of condition-level average adherence{note}\n"
            f"{n_adherent} of {len(by_condition)} conditions meet the {adherence_threshold}% threshold"
        ),
        xlabel="Average MEPS adherence ratio by condition (%)",
        ylabel="Number of conditions (ICD10CDX)",
        filename=f"condition_adherence_histogram_{tag}.png",
    )
    n_adherent_tc1 = (by_tc1s1["meps_adherence_ratio"] >= adherence_threshold).sum()
    _plot_adherence_histogram(
        ctx,
        by_tc1s1["meps_adherence_ratio"],
        title=(
            f"Distribution of TC1S1-level average adherence{note}\n"
            f"{n_adherent_tc1} of {len(by_tc1s1)} drug subclasses meet the {adherence_threshold}% threshold"
        ),
        xlabel="Average MEPS adherence ratio by TC1S1 drug subclass (%)",
        ylabel="Number of TC1S1 drug subclasses",
        filename=f"tc1s1_adherence_histogram_{tag}.png",
    )
    _save_table(
        ctx,
        by_condition[by_condition["meps_adherence_ratio"] >= adherence_threshold].sort_values(
            "meps_adherence_ratio", ascending=False
        ),
        f"adherent_conditions_{tag}.xlsx",
    )
    _save_table(
        ctx,
        by_tc1s1[by_tc1s1["meps_adherence_ratio"] >= adherence_threshold].sort_values(
            "meps_adherence_ratio", ascending=False
        ),
        f"adherent_tc1s1_{tag}.xlsx",
    )
    _save_table(
        ctx,
        by_tc1s1[by_tc1s1["meps_adherence_ratio"] < low_threshold].sort_values("meps_adherence_ratio"),
        f"low_tc1s1_{tag}.xlsx",
    )
    _save_table(
        ctx,
        by_condition[by_condition["meps_adherence_ratio"] < low_threshold].sort_values("meps_adherence_ratio"),
        f"low_conditions_{tag}.xlsx",
    )


def run_exports(year: int, meps_dir: str | Path | None = None, write_decisions_log: bool = True) -> None:
    """Run ``build()`` and write notebook-style tables/graphs to ``output/<year>/``."""
    import time

    import matplotlib

    matplotlib.use("Agg")

    if year not in YEAR_FILES:
        raise ValueError(f"year must be one of {sorted(YEAR_FILES)}; got {year}")

    tables_dir, graphs_dir = output_dirs(year)
    tables_dir.mkdir(parents=True, exist_ok=True)
    graphs_dir.mkdir(parents=True, exist_ok=True)

    ctx = ExportCtx(
        year=year,
        tables_dir=tables_dir,
        graphs_dir=graphs_dir,
        rx_p=rx_prefix(year),
    )
    t0 = time.perf_counter()
    dir_ = resolve_meps_dir() if meps_dir is None else Path(meps_dir)

    ctx.log("=" * 60)
    ctx.log(f"{year} MEPS clean pipeline  (via clean_meps.build)")
    ctx.log("=" * 60)
    ctx.info(f"MEPS data dir : {dir_}")
    ctx.info(f"Output dir    : {tables_dir.parent}")

    ctx.step("Build flat-365 grouped merge (all drugs, chronic conditions)")
    df_flat, bridge_flat, _log_flat = build(
        year, meps_dir=dir_, drug_chronic_only=False, pstats_denominator=False
    )
    ctx.info(f"{len(df_flat):,} person-drug pairs")
    _save_table(ctx, df_flat, "grouped_merge_df_flat365.xlsx")
    _save_table(
        ctx,
        pd.Series(
            {
                "unique_icd10": df_flat["ICD10CDX"].nunique(),
                "unique_tc1": df_flat["TC1"].nunique(),
                "unique_rxname": df_flat["RXNAME"].nunique(),
                "unique_tc1s1": df_flat["TC1S1"].nunique(),
            },
            name="count",
        ).to_frame(),
        "summary_counts_flat365.xlsx",
    )
    by_class = (
        df_flat.groupby(["TC1", "TC1S1"], as_index=False)
        .agg(meps_adherence_ratio=("meps_adherence_ratio", "mean"))
        .sort_values("meps_adherence_ratio", ascending=False)
    )
    _save_table(ctx, by_class.head(10), f"{ctx.rx_p}_high.xlsx")
    _save_table(ctx, by_class.tail(10).sort_values("meps_adherence_ratio"), f"{ctx.rx_p}_low.xlsx")

    ctx.step("Flat-365 condition/TC1S1 histograms and low-adherence tables")
    _export_summaries(ctx, df_flat, bridge_flat, tag="flat365")

    ctx.step("Build PSTATS-based grouped merge (all drugs)")
    df_pstats, bridge_pstats, _log_pstats = build(
        year, meps_dir=dir_, drug_chronic_only=False, pstats_denominator=True
    )
    ctx.info(f"{len(df_pstats):,} person-drug pairs")
    _save_table(ctx, df_pstats, "grouped_merge_df_pstats.xlsx")
    if "participation_type" in df_pstats.columns:
        _save_table(
            ctx,
            df_pstats["participation_type"].value_counts().to_frame("persons"),
            "participation_type_counts.xlsx",
        )
        ref_cols = [c for c in df_pstats.columns if c.startswith("ref_start_") or c.startswith("ref_end_")]
        sample_cols = ["DUPERSID", "total_days_supply", "participation_type", "coverage_notes"] + ref_cols
        _save_table(ctx, df_pstats[sample_cols].drop_duplicates("DUPERSID").head(10), "reference_days_sample.xlsx")
    _save_table(ctx, df_pstats["total_days_supply"].describe().to_frame(), "pstats_total_days_supply_summary.xlsx")
    _save_table(ctx, df_pstats["meps_adherence_ratio"].describe().to_frame(), "pstats_adherence_ratio_summary.xlsx")
    if "participation_type" in df_pstats.columns:
        compare = (
            df_pstats[["DUPERSID", "participation_type", "total_days_supply"]]
            .drop_duplicates("DUPERSID")
            .groupby("participation_type")
            .agg(
                persons=("DUPERSID", "count"),
                mean_days=("total_days_supply", "mean"),
                min_days=("total_days_supply", "min"),
                max_days=("total_days_supply", "max"),
            )
        )
        _save_table(ctx, compare.reset_index(), "pstats_participation_compare.xlsx")

    ctx.step("Build final frame (chronic drugs + PSTATS denominator)")
    df_final, bridge_final, run_log = build(year, meps_dir=dir_)
    ctx.info(
        f"{len(df_final):,} person-drug pairs, "
        f"{df_final['DUPERSID'].nunique():,} patients"
    )
    _save_table(ctx, df_final, "new_grouped_merge_df_chronic_drugs.xlsx")
    # Fast cache for all-years merge / Streamlit
    pq = ctx.tables_dir / "new_grouped_merge_df_chronic_drugs.parquet"
    _stringify_age_columns(df_final).to_parquet(pq, index=False)
    ctx.info(f"saved table -> output/{ctx.year}/tables/{pq.name}  ({len(df_final):,} rows)")

    # Patient-drug × condition bridge (chronic-drugs scope). Same grain as the
    # bridge returned by build(); consumers do bridge.merge(patient_drug, on=
    # ["DUPERSID","DRUGIDX"]) for condition-level rollups.
    bridge_pq = ctx.tables_dir / "patient_drug_condition_bridge.parquet"
    bridge_final.to_parquet(bridge_pq, index=False)
    ctx.info(
        f"saved table -> output/{ctx.year}/tables/{bridge_pq.name}  "
        f"({len(bridge_final):,} rows)"
    )
    _save_table(ctx, bridge_final, "patient_drug_condition_bridge.xlsx")
    _save_table(
        ctx,
        pd.Series(
            {
                "unique_icd10": df_final["ICD10CDX"].nunique(),
                "unique_tc1": df_final["TC1"].nunique(),
                "unique_rxname": df_final["RXNAME"].nunique(),
                "unique_tc1s1": df_final["TC1S1"].nunique(),
            },
            name="count",
        ).to_frame(),
        "summary_counts.xlsx",
    )

    ctx.step("Export unique RXNAME list from raw fills")
    rx_cols = _rx_cols(year)
    rx_raw = pd.read_excel(dir_ / YEAR_FILES[year]["rx"], engine="calamine", usecols=rx_cols)
    rx_raw = rx_raw[(rx_raw["RXDAYSUP"] > 0) & (rx_raw["RXDAYSUP"] < 990) & (rx_raw["RXBEGYRX"] > 0)]
    unique_rx = rx_raw[["RXNAME"]].drop_duplicates().sort_values("RXNAME").reset_index(drop=True)
    _save_table(ctx, unique_rx, "unique_rxname.csv")
    ctx.info(f"{len(unique_rx):,} unique drug names")

    ctx.step("Chronic-drug summaries, rankings, and histograms")
    by_drug_class = (
        df_final.groupby(["TC1", "TC1S1"], as_index=False)
        .agg(meps_adherence_ratio=("meps_adherence_ratio", "mean"))
        .sort_values("meps_adherence_ratio", ascending=False)
    )
    _save_table(ctx, by_drug_class.head(10), f"{ctx.rx_p}_high_chronic_drugs.xlsx")
    _save_table(ctx, by_drug_class.tail(10).sort_values("meps_adherence_ratio"), f"{ctx.rx_p}_low_chronic_drugs.xlsx")
    _export_summaries(ctx, df_final, bridge_final, tag="chronic_drugs", title_note="chronic drugs only")

    ctx.step("Top / bottom patient and patient-drug adherence (PSTATS denominator)")
    valid = df_final[df_final["total_days_supply"] > 0]
    patient_adherence = (
        valid.groupby("DUPERSID", as_index=False)
        .agg(
            total_valid_days=("total_valid_days", "sum"),
            total_days_supply=("total_days_supply", "sum"),
            drug_count=("DRUGIDX", "count"),
        )
    )
    patient_adherence["meps_adherence_ratio"] = (
        patient_adherence["total_valid_days"] / patient_adherence["total_days_supply"] * 100
    )
    _save_table(ctx, patient_adherence.nlargest(10, "meps_adherence_ratio"), "top_10_patients_high.xlsx")
    _save_table(ctx, patient_adherence.nsmallest(10, "meps_adherence_ratio"), "top_10_patients_low.xlsx")
    _save_table(ctx, valid.nlargest(10, "meps_adherence_ratio"), "top_10_patient_drug_high.xlsx")
    _save_table(ctx, valid.nsmallest(10, "meps_adherence_ratio"), "top_10_patient_drug_low.xlsx")

    ctx.step("Top 20 least-adherent condition/drug combinations")
    cond_view_final = _condition_view(df_final, bridge_final)
    top_20 = (
        cond_view_final.dropna(subset=["meps_adherence_ratio"])
        .groupby(["ICD10CDX", "ICD10CDX_LABEL", "RXNAME"], as_index=False)
        .agg(
            meps_adherence_ratio=("meps_adherence_ratio", "mean"),
            patient_drug_pairs=("DUPERSID", "count"),
        )
        .nsmallest(20, "meps_adherence_ratio")
        .sort_values("meps_adherence_ratio")
        .reset_index(drop=True)
    )
    _save_table(ctx, top_20, "top_20_least_adherent_chronic_drugs.xlsx")

    ctx.step("Write decisions log")
    if write_decisions_log:
        log_path = write_log(run_log)
        ctx.info(f"decisions log -> {log_path.name}")

    elapsed = time.perf_counter() - t0
    ctx.log("\n" + "=" * 60)
    ctx.log(f"Done in {elapsed / 60:.1f} min")
    ctx.log(f"Tables : {tables_dir}")
    ctx.log(f"Graphs : {graphs_dir}")
    ctx.log("=" * 60)

    # Keep the cross-year merge in sync whenever any single year is exported.
    ctx.step("Refresh all-years merged chronic-drug table")
    merged_path = export_merged_all_years(meps_dir=dir_)
    if merged_path is not None:
        ctx.info(f"all-years merge -> {merged_path}")


def enrich_model_df_all_years(meps_dir: str | Path | None = None) -> Path | None:
    """Attach paper enrichments onto ``model_df_all_years.parquet`` (Step R).

    Adds ``MARRYXX``, ``REGIONXX``, ``EDUCYR``, ``RACETHX``, and medication_*
    columns in place. No-op (returns None) when the panel parquet is missing.
    """
    import time

    dir_ = Path(meps_dir) if meps_dir is not None else resolve_meps_dir()
    tables_dir, _ = all_years_output_dirs()
    panel_path = tables_dir / MODEL_DF_ALL_YEARS
    if not panel_path.exists():
        print(f"[enrich] skip: {panel_path} not found")
        return None

    model_df = pd.read_parquet(panel_path)
    n_before = len(model_df)
    model_df = model_df.drop(columns=PAPER_ENRICH_COLS, errors="ignore")

    demo_parts: list[pd.DataFrame] = []
    med_parts: list[pd.DataFrame] = []
    for y in sorted(YEAR_FILES):
        t0 = time.perf_counter()
        print(f"[enrich] [{y}] loading person + Rx enrichments …")
        demo_parts.append(load_demo_extra_year(y, meps_dir=dir_))
        med_parts.append(load_med_extra_year(y, meps_dir=dir_))
        print(f"[enrich] [{y}] done in {time.perf_counter() - t0:.1f}s")

    demo_all = pd.concat(demo_parts, ignore_index=True)
    med_all = pd.concat(med_parts, ignore_index=True)
    med_all["medication_freq_bin"] = med_all["medication_freq"].map(_medication_freq_bin)

    model_df = model_df.merge(demo_all, on=["DUPERSID", "YEAR"], how="left")
    model_df = model_df.merge(med_all, on=["DUPERSID", "YEAR"], how="left")

    if len(model_df) != n_before:
        raise RuntimeError(
            f"row count changed after enrich: {n_before} → {len(model_df)}"
        )
    if int(model_df.duplicated(["DUPERSID", "YEAR"]).sum()) != 0:
        raise RuntimeError("duplicate (DUPERSID, YEAR) after enrich")

    model_df = _stringify_age_columns(model_df)
    model_df.to_parquet(panel_path, index=False)
    print(f"[enrich] wrote {panel_path} ({len(model_df):,} rows, {model_df.shape[1]} cols)")
    for c in PAPER_ENRICH_COLS:
        nn = int(model_df[c].notna().sum())
        print(f"[enrich]   {c}: {nn:,} / {len(model_df):,} ({100 * nn / len(model_df):.1f}%)")
    return panel_path


def enrich_year_chronic_frames(meps_dir: str | Path | None = None) -> list[Path]:
    """Patch per-year chronic-drug parquets with demos + person–drug medication_*.

    Use when year exports predate the paper enrichments so Streamlit sidebar
    filters work without a full ``run_exports`` rebuild.
    """
    import time

    dir_ = Path(meps_dir) if meps_dir is not None else resolve_meps_dir()
    written: list[Path] = []

    for year in sorted(YEAR_FILES):
        tables = output_dirs(year)[0]
        pq = tables / "new_grouped_merge_df_chronic_drugs.parquet"
        df = _read_year_chronic_frame(year)
        if df is None:
            print(f"[enrich-years] skip {year}: no chronic-drug export")
            continue

        t0 = time.perf_counter()
        print(f"[enrich-years] [{year}] attaching demos + medication_* …")
        n_before = len(df)
        df = df.drop(columns=PAPER_ENRICH_COLS, errors="ignore")

        demo = load_demo_extra_year(year, meps_dir=dir_).drop(columns=["YEAR"])
        df = df.merge(demo, on="DUPERSID", how="left")

        # Person–drug medication features from Rx (DRUGIDX grain)
        rx_path = dir_ / YEAR_FILES[year]["rx"]
        rx = pd.read_excel(
            rx_path,
            engine="calamine",
            usecols=[
                "DUPERSID", "DRUGIDX", "RXSTRENG", "RXSTRUNT",
                "RXQUANTY", "RXFORM", "RXDAYSUP",
            ],
        )
        med = _medication_features_from_fills(rx)
        if not med.empty and "DRUGIDX" in df.columns:
            df = df.merge(med, on=["DUPERSID", "DRUGIDX"], how="left")

        if len(df) != n_before:
            raise RuntimeError(
                f"{year}: row count changed after enrich: {n_before} → {len(df)}"
            )

        tables.mkdir(parents=True, exist_ok=True)
        _stringify_age_columns(df).to_parquet(pq, index=False)
        written.append(pq)
        print(
            f"[enrich-years] [{year}] wrote {pq.name} "
            f"({len(df):,} rows) in {time.perf_counter() - t0:.1f}s"
        )

    return written


def enrich_paper_columns(meps_dir: str | Path | None = None) -> dict[str, Path | None]:
    """Enrich year frames, all-years cache, and model_df with paper columns."""
    enrich_year_chronic_frames(meps_dir=meps_dir)
    merged = export_merged_all_years(meps_dir=meps_dir)
    model_path = enrich_model_df_all_years(meps_dir=meps_dir)
    return {"all_years": merged, "model_df": model_path}


# ---------------------------------------------------------------------------
# Legacy helper (kept for backwards compatibility with any code that imported
# clean_h248a from this module).
# ---------------------------------------------------------------------------

SENTINELS = [-1, -7, -8, -9, -15]


def clean_h248a(df: pd.DataFrame) -> pd.DataFrame:
    """Sentinel-mask h248a in place — pre-existing helper, unchanged.

    New code should call ``build(year)`` instead.
    """
    df = df.copy()
    if "DIABEQUIP" in df.columns:
        df = df[df["DIABEQUIP"] != 1].copy()
    for c in ("RXDAYSUP", "RXQUANTY", "RXSF23X", "RXMR23X"):
        if c in df.columns:
            df.loc[df[c].isin(SENTINELS), c] = np.nan
    if "RXDAYSUP" in df.columns:
        df.loc[df["RXDAYSUP"] == 999, "RXDAYSUP"] = np.nan
    if "RXDRGNAM" in df.columns:
        df = df[~df["RXDRGNAM"].astype(str).str.startswith("-")].copy()
    return df


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _cli(argv: list[str] | None = None) -> int:
    import argparse

    parser = argparse.ArgumentParser(
        prog="clean_meps",
        description="MEPS adherence build / export / all-years cache",
    )
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_cache = sub.add_parser(
        "cache-all-years",
        help="Build the all-years parquet cache + filter_options.json for Streamlit",
    )
    p_cache.add_argument(
        "--meps-dir",
        default=None,
        help="Optional override for the MEPS data directory",
    )

    p_export = sub.add_parser("export", help="Run full year export (tables + graphs)")
    p_export.add_argument("--year", type=int, required=True, choices=sorted(YEAR_FILES))
    p_export.add_argument("--meps-dir", default=None)

    p_enrich = sub.add_parser(
        "enrich-paper-cols",
        help=(
            "Attach MARRYXX/REGIONXX/EDUCYR/RACETHX + medication_* onto year "
            "chronic-drug caches, all-years merge, and model_df_all_years"
        ),
    )
    p_enrich.add_argument("--meps-dir", default=None)

    args = parser.parse_args(argv)

    if args.cmd == "cache-all-years":
        path = export_merged_all_years(meps_dir=args.meps_dir)
        return 0 if path is not None else 1

    if args.cmd == "export":
        run_exports(args.year, meps_dir=args.meps_dir)
        return 0

    if args.cmd == "enrich-paper-cols":
        out = enrich_paper_columns(meps_dir=args.meps_dir)
        ok = out.get("all_years") is not None
        return 0 if ok else 1

    return 1


if __name__ == "__main__":
    raise SystemExit(_cli())
